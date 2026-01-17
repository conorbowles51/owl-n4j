import csv
import json
from pathlib import Path
from typing import Dict, Optional, Callable

from openpyxl import load_workbook

from ingestion import ingest_document
from logging_utils import log_progress, log_error, log_warning


def extract_text_from_excel(path: Path) -> str:
    """
    Extract text from an Excel file.

    Preserves row and column context by labeling each cell's position.

    Args:
        path: Path to the Excel file

    Returns:
        Extracted text content with row/column context
    """
    workbook = load_workbook(str(path), data_only=True)
    chunks = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_lines = []

        for row_idx, row in enumerate(sheet.iter_rows(), start=1):
            row_cells = []
            for cell in row:
                if cell.value is not None:
                    col_letter = cell.column_letter
                    row_cells.append(f"[{col_letter}{row_idx}]: {cell.value}")
            if row_cells:
                sheet_lines.append(f"Row {row_idx}: " + " | ".join(row_cells))

        if sheet_lines:
            chunks.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(sheet_lines))

    workbook.close()
    return "\n\n".join(chunks)


def extract_text_from_csv(path: Path) -> str:
    """
    Extract text from a CSV file as JSON.

    Uses the CSV headers as keys for each row object.

    Args:
        path: Path to the CSV file

    Returns:
        JSON string representation of the CSV data
    """
    with open(path, newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        data = list(reader)

    return json.dumps(data, indent=2, default=str)

def ingest_excel_file(
    path: Path,
    case_id: str,
    log_callback: Optional[Callable[[str], None]] = None,
    profile_name: Optional[str] = None,
) -> Dict:
    if not case_id:
        raise ValueError("case_id is required for Excel file ingestion")

    doc_name = path.name
    extension = path.suffix.lower()

    log_progress(f"Extracting text from {path}", log_callback)

    try:
        if extension == ".csv":
            text = extract_text_from_csv(path)
            source_type = "csv"
        elif extension in (".xls", ".xlsx"):
            text = extract_text_from_excel(path)
            source_type = "excel"
        else:
            return {"status": "error", "reason": f"Unsupported file extension: {extension}", "file": str(path)}
    except Exception as e:
        log_error(f"Failed to extract text from {path}: {e}", log_callback)
        return {"status": "error", "reason": str(e), "file": str(path)}

    if not text.strip():
        log_warning(f"No text extracted from {path}, skipping", log_callback)
        return {"status": "skipped", "reason": "no_text", "file": str(path)}

    doc_metadata = {
        "filename": path.name,
        "full_path": str(path.resolve()),
        "source_type": source_type,
    }

    if source_type == "excel":
        workbook = load_workbook(str(path), data_only=True)
        doc_metadata["sheet_count"] = len(workbook.sheetnames)
        workbook.close()

    return ingest_document(
        text=text,
        doc_name=doc_name,
        case_id=case_id,
        doc_metadata=doc_metadata,
        log_callback=log_callback,
        profile_name=profile_name,
    )