"""
Text Extractor Processor

Extracts text content from PDF, DOCX, XLSX, and plain text files.
Reuses existing project dependencies (pypdf, python-docx, openpyxl).
"""

from __future__ import annotations

import logging
from typing import Any, Dict, List

from services.triage_processors.base_processor import BaseTriageProcessor, ProcessingResult

logger = logging.getLogger(__name__)


class TextExtractorProcessor(BaseTriageProcessor):
    name = "text_extractor"
    display_name = "Text Extractor"
    description = "Extract text content from documents (PDF, DOCX, XLSX, TXT)"
    input_types = ["documents"]
    output_types = ["extracted_text"]
    requires_llm = False
    config_schema = {
        "max_pages": {"type": "integer", "default": 100, "description": "Max pages to extract from PDFs"},
        "max_chars": {"type": "integer", "default": 500000, "description": "Max characters per document"},
    }

    def process_file(
        self,
        file_path: str,
        file_info: Dict[str, Any],
        config: Dict[str, Any],
    ) -> List[ProcessingResult]:
        ext = file_info.get("extension", "").lower()
        max_pages = config.get("max_pages", 100)
        max_chars = config.get("max_chars", 500000)

        try:
            if ext == ".pdf":
                text = self._extract_pdf(file_path, max_pages)
            elif ext in (".docx", ".doc"):
                text = self._extract_docx(file_path)
            elif ext in (".xlsx", ".xls"):
                text = self._extract_xlsx(file_path)
            elif ext in (".txt", ".md", ".csv", ".rtf", ".log"):
                text = self._extract_text(file_path)
            else:
                # Try plain text fallback
                text = self._extract_text(file_path)

            if text:
                text = text[:max_chars]
                return [ProcessingResult(
                    source_path=file_path,
                    artifact_type="extracted_text",
                    content=text,
                    metadata={
                        "char_count": len(text),
                        "source_format": ext,
                    },
                )]
            return []

        except Exception as e:
            return [ProcessingResult(
                source_path=file_path,
                artifact_type="extracted_text",
                error=str(e),
            )]

    def _extract_pdf(self, path: str, max_pages: int) -> str:
        try:
            from pypdf import PdfReader
        except ImportError:
            from PyPDF2 import PdfReader
        reader = PdfReader(path)
        pages = min(len(reader.pages), max_pages)
        texts = []
        for i in range(pages):
            page_text = reader.pages[i].extract_text()
            if page_text:
                texts.append(page_text)
        return "\n\n".join(texts)

    def _extract_docx(self, path: str) -> str:
        from docx import Document
        doc = Document(path)
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())

    def _extract_xlsx(self, path: str) -> str:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        texts = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(max_row=1000, values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    texts.append("\t".join(cells))
        wb.close()
        return "\n".join(texts)

    def _extract_text(self, path: str) -> str:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            return f.read()
